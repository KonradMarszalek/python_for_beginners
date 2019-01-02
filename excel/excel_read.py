from openpyxl import load_workbook
wb = load_workbook('My Sheet.xlsx')
# print(wb.sheetnames)

ws = wb.active

name_label = ws['A1']
scores_label = ws['B1']

print(name_label.value)
print(scores_label.value)


names = ws['A2':'B4']
scores = ws['A2':'A4']

print(names)
# print(scores)
